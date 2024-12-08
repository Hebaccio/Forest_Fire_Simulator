import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.colors as mcolors
import random
from scipy.ndimage import gaussian_filter
import tkinter as tk
from threading import Thread, Event
from openpyxl import Workbook, load_workbook
import os
from itertools import product  # Import for generating all combinations of parameters

# Grid dimensions
rows, cols = 100, 100

# Environmental variables
humidity = 0.8
precipitation_strength = 0.6
precipitation_chance = 0.7
wind_strength = 0.2

# Simulation state
stop_simulation_event = Event()

# Variables to store the initial state for restart
initial_forest = None
initial_moisture_map = None
initial_burn_timers = None

def initialize_forest(rows, cols):
    forest = np.random.choice([1, 6], size=(rows, cols), p=[0.6, 0.4])
    moisture_map = np.random.rand(rows, cols)
    burn_timers = np.zeros((rows, cols), dtype=int)
    return forest, moisture_map, burn_timers

def add_rock_clusters(forest, num_clusters, max_cluster_size):
    for _ in range(num_clusters):
        cluster_row = random.randint(0, rows - 1)
        cluster_col = random.randint(0, cols - 1)
        cluster_size = random.randint(1, max_cluster_size)
        for i in range(cluster_size):
            for j in range(cluster_size):
                ni, nj = cluster_row + i, cluster_col + j
                if 0 <= ni < rows and 0 <= nj < cols:
                    forest[ni, nj] = 4
    return forest

def add_water_clusters(forest, probability, sigma, threshold):
    water_layer = np.random.rand(rows, cols) < probability
    water_layer = gaussian_filter(water_layer.astype(float), sigma=sigma) > threshold
    forest[water_layer] = 3
    return forest

def ignite_random_fire(forest, burn_timers):
    while True:
        random_row = random.randint(0, rows - 1)
        random_col = random.randint(0, cols - 1)
        if forest[random_row, random_col] in [1, 6]:
            forest[random_row, random_col] = 2
            burn_timers[random_row, random_col] = 10 if forest[random_row, random_col] == 1 else 3
            break
    return forest, burn_timers

def spread_fire(grid, moisture_map, burn_timers, drying_effect, rain_active):
    global humidity, precipitation_strength, wind_strength
    new_grid = grid.copy()
    new_moisture = moisture_map.copy()
    new_burn_timers = burn_timers.copy()

    for i in range(rows):
        for j in range(cols):
            if grid[i, j] == 2:  # Burning
                if rain_active and precipitation_strength > 0.6:
                    new_grid[i, j] = 5  # Burnt
                    continue

                new_burn_timers[i, j] -= 1
                if new_burn_timers[i, j] <= 0:
                    new_grid[i, j] = 5  # Burnt
                    continue

                for dy, dx in [(0, 1), (1, 0), (0, -1), (-1, 0), (-1, -1), (-1, 1), (1, -1), (1, 1)]:
                    ni, nj = i + dy, j + dx
                    if 0 <= ni < rows and 0 <= nj < cols and grid[ni, nj] in [1, 6]:
                        reduction_factor = ((humidity + precipitation_strength) * 10) * 5
                        increase_factor = ((wind_strength + drying_effect) * 10) * 3
                        ignition_chance = 0.65 - (reduction_factor / 100) + (increase_factor / 100)
                        ignition_chance = max(0, min(1, ignition_chance))

                        if random.random() < ignition_chance:
                            new_grid[ni, nj] = 2
                            new_burn_timers[ni, nj] = 8 if grid[ni, nj] == 1 else 8

    new_moisture = np.clip(new_moisture - drying_effect * wind_strength, 0, 1)
    return new_grid, new_moisture, new_burn_timers

# Modified function to save results to Excel and print progress
def save_to_excel(humidity, precipitation_strength, precipitation_chance, wind_strength, drying_effect, burned_cells, burned_percentage, steps_taken, current_simulation, total_combinations):
    filename = "ForestFireSimulation.xlsx"
    
    if not os.path.exists(filename):
        # Create a new Excel file if it doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.title = "Simulation Results"
        # Add headers
        ws.append(["Humidity", "Precipitation Strength", "Precipitation Chance", "Wind Strength", "Drying Effect",
                   "Total m² Burned", "% Burned", "Total Minutes Taken"])
        wb.save(filename)

    # Load the workbook
    wb = load_workbook(filename)
    ws = wb.active

    # Add the results
    ws.append([humidity, precipitation_strength, precipitation_chance, wind_strength, drying_effect,
               burned_cells, burned_percentage, steps_taken])
    wb.save(filename)
    print(f"Results saved to {filename} {current_simulation}/{total_combinations}")

# Modified function to save results with progress
def run_simulation_without_visuals(forest, moisture_map, burn_timers, drying_effect, current_simulation, total_combinations):
    global total_trees  # Use the global variable for total trees

    if total_trees == 0:  # Safety check in case there are no trees
        raise ValueError("No trees in the landscape to simulate burning.")

    steps_taken = 0

    while True:
        rain_active = random.random() < precipitation_chance
        steps_taken += 1

        if np.all(forest != 2):  # No burning cells
            print("Fire has stopped spreading.")
            break

        forest, moisture_map, burn_timers = spread_fire(forest, moisture_map, burn_timers, drying_effect, rain_active)

    burned_cells = np.sum(forest == 5)
    burned_percentage = (burned_cells / total_trees) * 100  # Divide by total trees

    # Save results to Excel with progress information
    save_to_excel(humidity, precipitation_strength, precipitation_chance, wind_strength, drying_effect,
                  burned_cells, burned_percentage, steps_taken, current_simulation, total_combinations)

def start_simulation_without_visuals():
    # Check if the landscape has been generated
    if initial_forest is None or initial_moisture_map is None or initial_burn_timers is None:
        raise ValueError("Landscape not generated. Please generate the landscape before running simulations.")
    
    global forest, moisture_map, burn_timers  # Use the saved initial landscape
    forest = initial_forest.copy()
    moisture_map = initial_moisture_map.copy()
    burn_timers = initial_burn_timers.copy()

    # Ignite a fire and start the simulation
    forest, burn_timers = ignite_random_fire(forest, burn_timers)
    drying_effect = wind_strength
    simulation_thread = Thread(target=run_simulation_without_visuals, args=(forest, moisture_map, burn_timers, drying_effect), daemon=True)
    simulation_thread.start()

# Modified function to run the simulation for all combinations
def run_simulation_with_all_combinations():
    global forest, moisture_map, burn_timers  # Declare global variables
    
    # Possible values for parameters
    values = [0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    
    # Calculate the total number of combinations
    total_combinations = len(values) ** 4
    current_simulation = 0  # Counter for the current simulation
    
    # Iterate through all combinations of parameters
    for humidity_val, precipitation_strength_val, precipitation_chance_val, wind_strength_val in product(values, repeat=4):
        current_simulation += 1  # Increment the simulation counter
        
        # Set the global parameters
        global humidity, precipitation_strength, precipitation_chance, wind_strength
        humidity = humidity_val
        precipitation_strength = precipitation_strength_val
        precipitation_chance = precipitation_chance_val
        wind_strength = wind_strength_val

        # Initialize the landscape
        forest = initial_forest.copy()
        moisture_map = initial_moisture_map.copy()
        burn_timers = initial_burn_timers.copy()

        # Ignite a fire
        forest, burn_timers = ignite_random_fire(forest, burn_timers)

        # Run the simulation
        drying_effect = wind_strength
        run_simulation_without_visuals(forest, moisture_map, burn_timers, drying_effect, current_simulation, total_combinations)

    print("Simulations for all parameter combinations completed.")

# Modified button function to trigger the automated simulations
def start_simulation_with_automation():
    # Check if the landscape has been generated
    if initial_forest is None or initial_moisture_map is None or initial_burn_timers is None:
        raise ValueError("Landscape not generated. Please generate the landscape before running simulations.")
    
    # Start the automated simulations in a new thread
    simulation_thread = Thread(target=run_simulation_with_all_combinations, daemon=True)
    simulation_thread.start()

# Modify main to use the new automated simulation
def main():
    global initial_forest, initial_moisture_map, initial_burn_timers, results_window, results_label
    global humidity, precipitation_strength, precipitation_chance, wind_strength
    global humidity_input, precipitation_strength_input, precipitation_chance_input, wind_strength_input

    root = tk.Tk()
    root.title("Fire Simulation")

    fig, ax = plt.subplots(figsize=(8, 8))
    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

    def generate_landscape():
        global forest, moisture_map, burn_timers
        global initial_forest, initial_moisture_map, initial_burn_timers, total_trees  # Add total_trees to track the count
        
        # Generate the landscape
        forest, moisture_map, burn_timers = initialize_forest(rows, cols)
        forest = add_rock_clusters(forest, num_clusters=50, max_cluster_size=5)
        forest = add_water_clusters(forest, probability=0.15, sigma=3, threshold=0.2)
        
        # Save the initial state
        initial_forest = forest.copy()
        initial_moisture_map = moisture_map.copy()
        initial_burn_timers = burn_timers.copy()
        
        # Count the number of trees (values 1 and 6)
        total_trees = np.sum((forest == 1) | (forest == 6))
        print(f"Total number of trees in the generated landscape: {total_trees}")
        
        # Visualize the generated landscape
        ax.clear()
        ax.imshow(forest, cmap=mcolors.ListedColormap(['#654321', 'green', 'red', 'blue', 'grey', '#3d251e', '#5fa15f']),
                norm=mcolors.BoundaryNorm([-0.5, 0.5, 1.5, 2.5, 3.5, 4.5, 5.5, 6.5], 7))
        ax.set_title("Generated Landscape")
        canvas.draw()

    forest, moisture_map, burn_timers = initialize_forest(rows, cols)
    generate_landscape()

    tk.Button(root, text="Generate Landscape", command=generate_landscape, bg="blue", fg="white").pack(side=tk.LEFT, padx=5, pady=5)
    tk.Button(root, text="Start All Simulations", command=start_simulation_with_automation, bg="green", fg="white").pack(side=tk.RIGHT, padx=5, pady=5)

    root.mainloop()

if __name__ == "__main__":
    main()